#!/usr/bin/env python3
"""Turn a Tally "Stock Summary" export into site/products.js.

The export's product names are free-form strings; this script classifies each
into a category, extracts filterable attributes (brand, capacity, star rating,
type, …), applies the selling margin, and writes a JavaScript catalogue.

IMPORTANT — only the marked-up SELLING price is written out. Raw purchase costs
never leave this machine, so nothing sensitive lands in the public repo.

Usage:
    python tools/generate_catalog.py STOCK_SHEET.xlsx [--margin 0.05] [--out site/products.js]

Requires: openpyxl  (pip install openpyxl)
"""
import argparse
import json
import os
import re
import sys
from collections import Counter

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

BRANDS = {
    "BLUE STAR": "Blue Star", "VOLTAS": "Voltas", "DAIKIN": "Daikin", "CARRIER": "Carrier",
    "GODREJ": "Godrej", "HAIER": "Haier", "WHIRL": "Whirlpool", "LLOYD": "Lloyd",
    "MITS": "Mitsubishi", "BAJAJ": "Bajaj", "ATOMBERG": "Atomberg", "CROMPTON": "Crompton",
    "USHA": "Usha", "FABER": "Faber", "MORPHY": "Morphy Richards", "SUNFLAME": "Sunflame",
    "SYMPHONY": "Symphony", "ORIENT": "Orient", "SINGER": "Singer", "VENUS": "Venus",
    "OSCAR": "Oscar", "EXIDE": "Exide", "ZEBRONICS": "Zebronics", "LEXTEL": "Lextel",
    "BPL": "BPL", "VESTAR": "Vestar", "LG": "LG", "SAM": "Samsung", "EF ": "Eureka Forbes",
}
TV_SIZES = ["24", "32", "40", "43", "50", "55", "65"]


def brand_of(u):
    u = re.sub(r"^\d+\s+", "", u)  # drop a leading model number ("104 LEXTEL ...")
    if u.startswith("BLUE STAR"):
        return "Blue Star"
    for key, val in BRANDS.items():
        if u.startswith(key):
            return val
    return u.split()[0].title()


def star(u):
    m = re.search(r"(\d)\s*\*", u)
    return f"{m.group(1)} Star" if m else None


def tonnage(u):
    m = re.search(r"(\d(?:\.\d)?)\s*TON", u)
    return f"{float(m.group(1)):g} Ton" if m else None


def litres(u):
    m = re.search(r"(\d{2,3})\s*L(?![A-Z])", u)
    return f"{m.group(1)} L" if m else None


def tv_size(u):
    m = re.search(r"(\d{2,3})\s*\"", u)
    if not m:
        m = re.search(r"LED\s*[A-Z]*-?(\d{2})", u)
    if not m:
        m = re.search(r"UA(\d{2})", u)
    if m:
        return f"{m.group(1)} inch"
    for s in TV_SIZES:
        if re.search(r"\b" + s, u):
            return f"{s} inch"
    return None


def categorize(u):
    """Return (category, attributes) or (None, None) to drop the row."""
    if "TALLY SOFTWARE" in u:
        return None, None
    if "INSTALATION KIT" in u or "INSTALLATION KIT" in u or u.strip() == "ODU STAND":
        return "Accessories", {"Brand": brand_of(u)}

    b = brand_of(u)

    # Whirlpool refrigerators carry no REF token ("215 ... ROY ... 3S")
    if b == "Whirlpool" and "ROY" in u:
        sm = re.search(r"(\d)\s*S\b", u)
        cap = re.search(r"\b(\d{3})\b", u)
        wtype = "Frost Free" if ("VITAMAGIC" in u or "INV" in u) else "Direct Cool"
        return "Refrigerator", {
            "Brand": b, "Type": wtype,
            "Capacity": f"{cap.group(1)} L" if cap else "—",
            "Star Rating": f"{sm.group(1)} Star" if sm else "—",
        }

    # Air conditioner: SAC = split, WAC = window, or anything tonnage-rated
    if re.search(r"\bSAC\b", u) or re.search(r"\bWAC\b", u) or tonnage(u):
        typ = "Window" if re.search(r"\bWAC\b", u) else "Split"
        drive = "Inverter" if "INV" in u else "Fixed Speed"
        cap, st = tonnage(u), star(u)
        if cap is None or st is None:  # model codes: Voltas 183=1.5T/3*, Godrej 18ITC5=1.5T/5*
            vm = re.search(r"\b(12|18|24)(?:[A-Z]{1,4})?([2-5])(?=\D|$)", u)
            if vm:
                cap = cap or {"12": "1 Ton", "18": "1.5 Ton", "24": "2 Ton"}[vm.group(1)]
                st = st or f"{vm.group(2)} Star"
        return "Air Conditioner", {
            "Brand": b, "Type": typ, "Capacity": cap or "—",
            "Star Rating": st or "—", "Compressor": drive,
        }

    if "COOLER" in u:
        return "Air Cooler", {"Brand": b}

    if re.search(r"\bREF\b", u) or "MINI BAR" in u:
        typ = ("Frost Free" if re.search(r"\bFF\b", u) else
               "Direct Cool" if re.search(r"\bDC\b", u) else
               "Mini Bar" if "MINI BAR" in u else "—")
        return "Refrigerator", {"Brand": b, "Type": typ, "Capacity": litres(u) or "—"}

    if re.search(r"\bWM\b", u) or "WASHING" in u:
        t = ("Front Load" if "FRONT LOAD" in u else "Top Load" if "TOP LOAD" in u else
             "Semi Automatic" if "SEMI" in u else "—")
        return "Washing Machine", {"Brand": b, "Type": t}

    if "LED" in u:  # every LED entry in this catalogue is a TV
        return "Television", {"Brand": b, "Screen Size": tv_size(u) or "—"}

    if "OTG" in u:
        return "Oven (OTG)", {"Brand": b, "Capacity": litres(u) or "—"}
    if re.search(r"M/W|M/O|\bMO\b|MICROWAVE|M O ", u):
        return "Microwave", {"Brand": b}

    if re.search(r"\bCF\b", u) or "CELING FAN" in u or "CEILING FAN" in u:
        return "Ceiling Fan", {"Brand": b}
    if "PADESTAL" in u or "PEDESTAL" in u:
        return "Pedestal Fan", {"Brand": b}
    if "EXH" in u or "BRISKAIR" in u or "CRISP AIR" in u:
        return "Exhaust Fan", {"Brand": b}

    if "WATER DISPENSER" in u:
        return "Water Dispenser", {"Brand": b}
    if "SURE CHAMP" in u or "PURIFIER" in u:
        return "Water Purifier", {"Brand": b}
    if re.search(r"\bWH\b", u) or "WATER HEATER" in u or "AQUAM" in u:
        return "Water Heater", {"Brand": b, "Capacity": litres(u) or "—"}

    if "KETTLE" in u:
        return "Electric Kettle", {"Brand": b, "Capacity": litres(u) or "—"}
    if "INDUCTION" in u:
        return "Induction Cooktop", {"Brand": b}
    if "COOKTOP" in u or "LPG" in u or re.search(r"\dBUR|BURNER", u):
        return "Gas Stove / Cooktop", {"Brand": b}
    if "HOOD" in u or "CHIMNEY" in u:
        return "Chimney", {"Brand": b}
    if "SANDWICH" in u:
        return "Sandwich Toaster", {"Brand": b}
    if "TOASTER" in u:
        return "Toaster", {"Brand": b}
    if "MIXER" in u:
        return "Mixer Grinder", {"Brand": b}
    if "COOKWARE" in u:
        return "Cookware", {"Brand": b}

    if "ROOM HEATER" in u or "HEAT CONVERTOR" in u or "HEATER" in u:
        return "Room Heater", {"Brand": b}

    if "PHONE" in u:
        return "Landline Phone", {"Brand": b}
    if "SPEAKER" in u:
        return "Speaker", {"Brand": b}
    if "EXIDE" in u or re.search(r"\(\d+AH\)", u) or "BATTERY" in u:
        return "Battery", {"Brand": b}

    return "Other", {"Brand": b}


def main():
    ap = argparse.ArgumentParser(description="Generate site/products.js from a Tally stock export.")
    ap.add_argument("xlsx", help="Path to the Tally Stock Summary .xlsx export")
    ap.add_argument("--margin", type=float, default=0.05, help="Selling margin over cost (default 0.05 = 5%%)")
    ap.add_argument("--sheet", default=None, help="Worksheet name (default: first sheet)")
    ap.add_argument("--out", default=None, help="Output path (default: <repo>/site/products.js)")
    ap.add_argument("--min-cost", type=float, default=10.0, help="Drop rows with cost below this (data glitches)")
    args = ap.parse_args()

    out = args.out or os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "site", "products.js")

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.worksheets[0]

    products, dropped = [], []
    for r in range(1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        qty = ws.cell(r, 2).value
        rate = ws.cell(r, 3).value
        if not name or not isinstance(rate, (int, float)):
            continue  # header / total / blank rows have no numeric rate
        name = str(name).strip()
        if name.lower() == "grand total":
            continue
        try:
            q = int(qty)
        except (TypeError, ValueError):
            q = 0
        cost = float(rate)
        if q <= 0 or cost < args.min_cost:
            dropped.append((name, "out-of-stock / low-cost glitch"))
            continue
        cat, attrs = categorize(name.upper())
        if cat is None:
            dropped.append((name, "non-retail"))
            continue
        products.append({
            "name": name,
            "category": cat,
            "attributes": attrs,
            "stock": q,
            "price": int(round(cost * (1 + args.margin))),
        })

    header = (
        "// Auto-generated by tools/generate_catalog.py — do not edit by hand.\n"
        f"// Selling price = last purchase cost x {1 + args.margin:g} (a {args.margin * 100:g}% margin).\n"
        "// Raw cost prices are intentionally NOT included in this file.\n"
    )
    with open(out, "w", encoding="utf-8") as f:
        f.write(header + "const PRODUCTS = " + json.dumps(products, ensure_ascii=False, indent=2) + ";\n")

    cats = Counter(p["category"] for p in products)
    print(f"Wrote {len(products)} products to {out}  (margin {args.margin * 100:g}%)")
    for c, n in sorted(cats.items(), key=lambda x: -x[1]):
        print(f"  {n:3d}  {c}")
    if dropped:
        print(f"\nDropped {len(dropped)} row(s):")
        for name, why in dropped:
            print(f"  - {name}  ({why})")


if __name__ == "__main__":
    main()
